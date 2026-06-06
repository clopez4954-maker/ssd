import io
import csv
from django.views.generic import TemplateView
from django.db.models import Count, Avg, Max, Min, Q
from django.http import HttpResponse
from django.utils import timezone

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from apps.usuarios.decoradores import EvaluadorRequeridoMixin
from apps.usuarios.models import LogAccion
from apps.postulantes.models import Postulante
from apps.evaluaciones.models import Evaluacion


# ──────────────────────────────────────────────
# Dashboard estadístico
# ──────────────────────────────────────────────

class DashboardReportesView(EvaluadorRequeridoMixin, TemplateView):
    template_name = 'reportes/dashboard.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        # Totales generales
        ctx['total_postulantes'] = Postulante.objects.count()
        ctx['total_evaluados'] = Evaluacion.objects.exclude(estado='pendiente').count()
        ctx['total_aprobados'] = Evaluacion.objects.filter(estado='aprobado').count()
        ctx['total_rechazados'] = Evaluacion.objects.filter(estado='rechazado').count()
        ctx['total_pendientes'] = Evaluacion.objects.filter(estado='pendiente').count()

        # Promedio de puntajes
        stats = Evaluacion.objects.exclude(estado='pendiente').aggregate(
            avg=Avg('puntaje_total'),
            maximo=Max('puntaje_total'),
            minimo=Min('puntaje_total'),
        )
        ctx['puntaje_promedio'] = round(stats['avg'] or 0, 2)
        ctx['puntaje_max'] = stats['maximo'] or 0
        ctx['puntaje_min'] = stats['minimo'] or 0

        # Distribución por estrato (para Chart.js)
        estrato_data = []
        for estrato in range(1, 7):
            total = Postulante.objects.filter(
                datos_socioeconomicos__estrato=estrato
            ).count()
            aprobados = Evaluacion.objects.filter(
                postulante__datos_socioeconomicos__estrato=estrato,
                estado='aprobado',
            ).count()
            estrato_data.append({
                'estrato': estrato,
                'total': total,
                'aprobados': aprobados,
            })
        ctx['estrato_data'] = estrato_data

        # Distribución de estados (para gráfica de donut)
        ctx['estado_data'] = {
            'labels': ['Pendiente', 'Aprobado', 'Rechazado'],
            'values': [
                ctx['total_pendientes'],
                ctx['total_aprobados'],
                ctx['total_rechazados'],
            ],
        }

        # Top 10 postulantes por puntaje
        ctx['top_postulantes'] = Evaluacion.objects.filter(
            estado='aprobado'
        ).select_related('postulante').order_by('-puntaje_total')[:10]

        return ctx


# ──────────────────────────────────────────────
# Exportar CSV
# ──────────────────────────────────────────────

class ExportarCSVView(EvaluadorRequeridoMixin, TemplateView):
    def get(self, request, *args, **kwargs):
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = (
            f'attachment; filename="postulantes_{timezone.now():%Y%m%d_%H%M}.csv"'
        )
        response.write('\ufeff')  # BOM para Excel en español

        writer = csv.writer(response)
        writer.writerow([
            'Cédula', 'Nombre Completo', 'Email', 'Institución', 'Carrera',
            'Semestre', 'Promedio', 'Créditos Aprobados',
            'Ingreso Familiar', 'Estrato',
            'Puntaje Académico', 'Puntaje Socioeconómico', 'Puntaje Total',
            'Estado', 'Fecha Evaluación',
        ])

        postulantes = Postulante.objects.select_related(
            'user', 'datos_academicos', 'datos_socioeconomicos', 'evaluacion'
        ).order_by('-evaluacion__puntaje_total')

        for p in postulantes:
            acad = getattr(p, 'datos_academicos', None)
            socio = getattr(p, 'datos_socioeconomicos', None)
            ev = getattr(p, 'evaluacion', None)
            writer.writerow([
                p.cedula,
                p.nombre_completo,
                p.user.email,
                acad.institucion if acad else '',
                acad.carrera if acad else '',
                acad.semestre if acad else '',
                acad.promedio if acad else '',
                acad.creditos_aprobados if acad else '',
                socio.ingreso_familiar if socio else '',
                socio.estrato if socio else '',
                ev.puntaje_academico if ev else '',
                ev.puntaje_socioeconomico if ev else '',
                ev.puntaje_total if ev else '',
                ev.get_estado_display() if ev else 'Pendiente',
                ev.fecha_evaluacion.strftime('%d/%m/%Y %H:%M') if ev and ev.fecha_evaluacion else '',
            ])

        LogAccion.objects.create(usuario=request.user, accion='exportar',
                                 detalles={'formato': 'CSV'})
        return response


# ──────────────────────────────────────────────
# Exportar Excel
# ──────────────────────────────────────────────

class ExportarExcelView(EvaluadorRequeridoMixin, TemplateView):
    def get(self, request, *args, **kwargs):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Postulantes'

        # ── Estilos ────────────────────────────────
        azul = '1e3a5f'
        verde = '27ae60'
        rojo = 'e74c3c'
        amarillo = 'f39c12'

        header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill('solid', fgColor=azul)
        center = Alignment(horizontal='center', vertical='center')
        thin = Side(style='thin', color='CCCCCC')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        headers = [
            ('Cédula', 16), ('Nombre Completo', 30), ('Email', 28),
            ('Institución', 25), ('Carrera', 25), ('Semestre', 10),
            ('Promedio', 12), ('Créditos', 10),
            ('Ingreso Familiar', 18), ('Estrato', 10),
            ('Pts. Académico', 15), ('Pts. Socioecon.', 15), ('Pts. Total', 12),
            ('Estado', 14), ('Fecha Evaluación', 20),
        ]

        # ── Cabecera ───────────────────────────────
        for col_idx, (header, width) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.row_dimensions[1].height = 25

        # ── Datos ──────────────────────────────────
        postulantes = Postulante.objects.select_related(
            'user', 'datos_academicos', 'datos_socioeconomicos', 'evaluacion'
        ).order_by('-evaluacion__puntaje_total')

        fill_aprobado = PatternFill('solid', fgColor='d5f5e3')
        fill_rechazado = PatternFill('solid', fgColor='fadbd8')
        fill_pendiente = PatternFill('solid', fgColor='fef9e7')

        for row_idx, p in enumerate(postulantes, 2):
            acad = getattr(p, 'datos_academicos', None)
            socio = getattr(p, 'datos_socioeconomicos', None)
            ev = getattr(p, 'evaluacion', None)

            estado = ev.get_estado_display() if ev else 'Pendiente'
            fila_fill = (
                fill_aprobado if ev and ev.estado == 'aprobado'
                else fill_rechazado if ev and ev.estado == 'rechazado'
                else fill_pendiente
            )

            row_data = [
                p.cedula, p.nombre_completo, p.user.email,
                acad.institucion if acad else '', acad.carrera if acad else '',
                acad.semestre if acad else '', float(acad.promedio) if acad else '',
                acad.creditos_aprobados if acad else '',
                float(socio.ingreso_familiar) if socio else '', socio.estrato if socio else '',
                float(ev.puntaje_academico) if ev else '',
                float(ev.puntaje_socioeconomico) if ev else '',
                float(ev.puntaje_total) if ev else '',
                estado,
                ev.fecha_evaluacion.strftime('%d/%m/%Y %H:%M') if ev and ev.fecha_evaluacion else '',
            ]

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                cell.fill = fila_fill
                if col_idx in (7, 11, 12, 13):
                    cell.number_format = '#,##0.00'

        # ── Congelar cabecera ──────────────────────
        ws.freeze_panes = 'A2'

        # ── Hoja de estadísticas ───────────────────
        ws2 = wb.create_sheet('Estadísticas')
        ws2.append(['Métrica', 'Valor'])
        ws2.append(['Total postulantes', Postulante.objects.count()])
        ws2.append(['Aprobados', Evaluacion.objects.filter(estado='aprobado').count()])
        ws2.append(['Rechazados', Evaluacion.objects.filter(estado='rechazado').count()])
        ws2.append(['Pendientes', Evaluacion.objects.filter(estado='pendiente').count()])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = (
            f'attachment; filename="becas_{timezone.now():%Y%m%d_%H%M}.xlsx"'
        )
        LogAccion.objects.create(usuario=request.user, accion='exportar',
                                 detalles={'formato': 'Excel'})
        return response
